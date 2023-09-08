
from openpyxl import Workbook

##seleccion nuevo doc o cargar anterior
import openpyxl

decisiondocnuevaopsasada=(input("Desea hacer un nuevo doc para tomar datos? si o no")).lower().replace(" ","")

variable1dedesicion= True
while variable1dedesicion:
    if decisiondocnuevaopsasada == "si":
        print("Listo continue")
        variable1dedesicion= False
####seleccion nuevo doc o cargar anterior fin
        wb= Workbook()
        titulo_del_excel= str(input("Nombre del excel:"))
        titulodela_hoja= str(input("Nombre de la hoja de excel:"))
        ubicaciondeguardado= "D:\\users\\jeff\\desktop\\jeffaprende" + "\\" +titulo_del_excel + ".xlsx"

        wb["Sheet"].title= titulodela_hoja
        sh1= wb.active
        sh1["A1"].value = "genero"
        sh1["B1"].value= "edad"
        sh1["C1"].value= "vestimenta_sup"
        sh1["D1"].value= "talla_sup"
        sh1["E1"].value= "color_sup"
        sh1["F1"].value= "vestimenta_inf"
        sh1["G1"].value= "talla_inf"
        sh1["H1"].value= "color_inf"
        sh1["I1"].value= "marcas"
        sh1["J1"].value= "gustos parte sup"
        sh1["K1"].value= "nivel_adquisitivo"
        sh1["L1"].value= "tiene_guantes?"
        sh1["M1"].value= "que_accesorio_usa?"
        sh1["N1"].value = "contextura"

##seleccion nuevo doc o cargar anterior
    elif decisiondocnuevaopsasada == "no":
        variable1dedesicion=False

# partedecargararchivo viejo inicio
        nombredelarchivopasado = str(input("ingrese el nombre del archivo que desea continuar"))
        #rutadelarchivopasado= input("ingrese sin comillas la ruta del archivo que desea continuar")
        wb = openpyxl.load_workbook("D:\\users\\jeff\\desktop\\jeffaprende" + "\\" + nombredelarchivopasado + ".xlsx")
        #wb = openpyxl.load_workbook(rutadelarchivopasado)
        print((wb))
        sheets = wb.sheetnames
        print("Nombres de las hojas del doc", sheets)
        hojaquedeseacontinuar = input("Escriba el nombre de la hoja que desea continuar")
        sh1 = wb[hojaquedeseacontinuar]
        rowsh1 = sh1.max_row
        print("Hay", rowsh1 - 1, "encuestados")
        #ubicaciondeguardado=rutadelarchivopasado
        ubicaciondeguardado = "D:\\users\\jeff\\desktop\\jeffaprende" + "\\" + nombredelarchivopasado + ".xlsx"
        # parte de cargar archivo viejo final

    else:
        decisiondocnuevaopsasada = (input("Desea hacer un nuevo doc para tomar datos? si o no")).lower().replace(" ", "")
####seleccion nuevo doc o cargar anterior fin



vestimenta_sup_opc = ["c =Camiseta", "b =Bividi", "t =top","ct = camiseta top" ,"s =Sudadera licra","cl =camisa licra","e =enterizo"]

vestimenta_inf_opc = ["p =pantaloneta","j =jogger", "ls =licra-short","lg =leggins", "s =shorts(bodys)", "e =enterizo","f =falda deportiva"]

tallas_opc=["XS","S","M","L","XL"]

marcas_deportivas= ["nike","adidas","reebok","puma","under armour","otro"]

gustos_opc=["A =Apretado","C =Comodo"]

nivel_adquisitivo_opc= ["b= BAJO", "m= MEDIO", "a= ALTO"]

edad_opc= "cuantos años de 15 a 50?"

guantes= ["Si", "No"]

accesorios_opc= ["mu =muñequera","r =rodillera", "ma =mangas", "c =coderas","g =gorra","no"]

contextura_opc= ["delgado","pepudo","ancho","bc","no"]

colores_opc= ["negro","azul","cafe","gris","verde","naranja","rosa","purpura","morado","conchevino","beige","plateado","rojo","blanco","amarillo","celeste","turquesa","dorado"]
#
respuestas= ["m","f","c","b","t","s","e","cl","p","ct","j","ls","lg","s","xs","s","m","l","xl","a","c","b","m","a","si","no","mu","r", "ma","g","c","no","negro","azul","cafe","gris","verde","naranja","rosa","purpura","morado","conchevino","beige","plateado","rojo","blanco","amarillo","celeste","turquesa","dorado","nike","adidas","reebok","puma","underarmour","otro","delgado","pepudo","ancho","bc","no",15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,0,"0"]
#

rowsh1= sh1.max_row
colsht1= sh1.max_column


#Parte1
numeros_de_encuestados= int(input("Cuantas personas encuestaras?")) +rowsh1 -1


contador_enc= 0 + rowsh1 -1



while contador_enc <= numeros_de_encuestados-1:
    contador_enc= contador_enc + 1
    print(contador_enc)

    #inicio pt2
    for i in range(1, colsht1 + 1):
        clasificacion = sh1.cell(1, i).value
        print(clasificacion)
        if clasificacion == "vestimenta_sup":
            print(vestimenta_sup_opc)
        elif clasificacion == "vestimenta_inf":
            print(vestimenta_inf_opc)
        elif clasificacion == "talla_sup":
            print(tallas_opc)
        elif clasificacion == "talla_inf":
            print(tallas_opc)
        elif clasificacion == "gustos parte sup":
            print(gustos_opc)
        elif clasificacion == "Nivel_Adquisitivo":
            print(nivel_adquisitivo_opc)
        elif clasificacion == "que_accesorio_usa?":
            print(accesorios_opc)
        elif clasificacion == "nivel_adquisitivo":
            print(nivel_adquisitivo_opc)
        elif clasificacion == "color_sup":
            print(colores_opc)
        elif clasificacion == "color_inf":
            print(colores_opc)
        elif clasificacion == "marcas":
            print(marcas_deportivas)
        elif clasificacion == "edad":
            print(edad_opc)
        elif clasificacion == "contextura":
            print(contextura_opc)

        valor_celda = input()
        if clasificacion == "edad":
            valor_celda= int(valor_celda)
        elif type(valor_celda) is str:
            valor_celda = valor_celda.lower().replace(" ","")


        #verificacionderespuesta inicio

        prueba_respuestas= not( valor_celda in respuestas)
        while prueba_respuestas:
            if prueba_respuestas is True:
                print("reponda correctamente por favor C:")
            valor_celda = input()
            prueba_respuestas = not (valor_celda in respuestas)
        #verificacionderespuesta fin

        sh1.cell(row=contador_enc + 1, column=i, value=valor_celda)
    #finalpt2

    if contador_enc  == numeros_de_encuestados:
        prgt_continua= str(input("Quiere seguir encuestando?").lower())
        comparacion= (prgt_continua=="si") or (prgt_continua=="no")

        variablecorrecionderespuesta= True
        while variablecorrecionderespuesta:

            if prgt_continua == "si":
                cuantosmas= int(input("Cuantos mas quieres encuestar? "))
                variablecorrecionderespuesta= False

                numeros_de_encuestados= numeros_de_encuestados+cuantosmas

            elif prgt_continua == "no":
                print("Listo lleno una hoja de datos con exito :D")
                variablecorrecionderespuesta = False
                print("Ubicacion de guardado:",ubicaciondeguardado)
            else:
                prgt_continua = str(input("Quiere seguir encuestando?").lower())




    wb.save(ubicaciondeguardado)




